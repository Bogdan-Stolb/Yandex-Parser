
import os, re, time, json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, NoSuchElementException

class YandexMapsParser:
    def __init__(self, headless=False, delay=1.5, debug=True):
        self.debug = debug
        options = webdriver.ChromeOptions()
        if headless:
            options.add_argument('--headless')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
        options.add_experimental_option("prefs", {
            "profile.default_content_setting_values.geolocation": 1,
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False
        })
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=options)
        self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        })
        self.wait = WebDriverWait(self.driver, 20)
        self.actions = ActionChains(self.driver)
        self.delay = delay
        
        if self.debug:
            print("[DEBUG] Браузер инициализирован")
    
    def _log(self, msg):
        if self.debug:
            print(f"[DEBUG] {msg}")
    
    def _safe_find(self, by, selector, timeout=10):
        try:
            return WebDriverWait(self.driver, timeout).until(EC.presence_of_element_located((by, selector)))
        except TimeoutException:
            self._log(f"Элемент не найден: {selector}")
            return None
        except Exception as e:
            self._log(f"Ошибка поиска {selector}: {str(e)[:50]}")
            return None
    
    def _safe_find_all(self, by, selector):
        try:
            return self.driver.find_elements(by, selector)
        except Exception as e:
            self._log(f"Ошибка поиска всех {selector}: {str(e)[:50]}")
            return []
    
    def _is_ad(self, card):
        try:
            html = (card.get_attribute('class') + ' ' + card.get_attribute('outerHTML')).lower()
            return any(x in html for x in ['promo', 'advertisement', 'business-card-promo', 'yandex-direct', '_advertisement'])
        except Exception as e:
            self._log(f"Ошибка проверки рекламы: {str(e)[:50]}")
            return False
    
    def _extract_yandex_link(self, card_or_driver, in_card_view=False):
        try:
            if in_card_view:
                link_el = card_or_driver.find_element(By.CSS_SELECTOR, "a.card-title-view__title-link[href^='/maps/org/']")
            else:
                link_el = card_or_driver.find_element(By.CSS_SELECTOR, "a.link-overlay[href^='/maps/org/']")
            href = link_el.get_attribute('href')
            if href and href.startswith('/maps/org/'):
                return f"https://yandex.ru{href.split('?')[0]}"
            elif href and href.startswith('https://'):
                return href.split('?')[0]
        except Exception as e:
            self._log(f"Ошибка извлечения ссылки: {str(e)[:50]}")
        return "N/A"
    
    def _extract_rating(self):
        try:
            rating_el = WebDriverWait(self.driver, 3).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.business-card-view span.business-rating-badge-view__rating-text"))
            )
            rating = rating_el.text.strip().replace(',', '.')
            if re.match(r'^[0-5](\.[0-9])?$', rating):
                return rating
        except:
            pass
        try:
            stars_container = self.driver.find_element(By.CSS_SELECTOR, "div.business-card-view div.business-rating-badge-view__stars")
            aria = stars_container.get_attribute('aria-label')
            if aria:
                match = re.search(r'Оценка\s*([\d,\.]+)', aria)
                if match:
                    return match.group(1).replace(',', '.')
        except Exception as e:
            self._log(f"Ошибка рейтинга: {str(e)[:50]}")
        return "N/A"
    
    def _parse_reviews(self):
        reviews = []
        try:
            review_items = self._safe_find_all(By.CSS_SELECTOR, "div.business-reviews-card-view__review")[:3]
            for item in review_items:
                try:
                    author = item.find_element(By.CSS_SELECTOR, "a.business-review-view__link span[itemprop='name']").text.strip()
                except:
                    author = "Аноним"
                try:
                    rating = item.find_element(By.CSS_SELECTOR, "meta[itemprop='ratingValue']").get_attribute('content')
                except:
                    rating = "N/A"
                try:
                    text_el = item.find_element(By.CSS_SELECTOR, "span.spoiler-view__text-container[data-original-size]")
                    text = text_el.text.strip()
                except:
                    try:
                        text = item.find_element(By.CSS_SELECTOR, "div.business-review-view__body").text.strip()
                    except:
                        text = ""
                try:
                    date = item.find_element(By.CSS_SELECTOR, "meta[itemprop='datePublished']").get_attribute('content')[:10]
                except:
                    try:
                        date = item.find_element(By.CSS_SELECTOR, "span.business-review-view__date span").text.strip()
                    except:
                        date = "N/A"
                reviews.append({'author': author, 'rating': rating, 'text': text, 'date': date})
        except Exception as e:
            self._log(f"Ошибка отзывов: {str(e)[:40]}")
        return reviews
    
    def _wait_for_search_input(self, timeout=15):
        """Ждём появления поисковой строки"""
        try:
            search_input = WebDriverWait(self.driver, timeout).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input.input__control"))
            )
            self._log("Поисковая строка найдена")
            return search_input
        except TimeoutException:
            self._log("Поисковая строка не найдена за 15 секунд")
            return None
        except Exception as e:
            self._log(f"Ошибка ожидания поисковой строки: {str(e)[:50]}")
            return None
    
    def _clear_and_type(self, search_input, text):
        """Очищаем поле и вводим текст"""
        try:
            search_input.click()
            time.sleep(0.2)
            search_input.clear()
            time.sleep(0.2)
            search_input.send_keys(Keys.CONTROL, 'a')
            search_input.send_keys(Keys.DELETE)
            time.sleep(0.1)
            search_input.send_keys(text)
            self._log(f"Введено: '{text}'")
            return True
        except Exception as e:
            self._log(f"Ошибка ввода текста: {str(e)[:50]}")
            return False
    
    def set_city(self, city_name):
        self._log(f"станавливаем город: {city_name}")
        
        try:
            self.driver.get("https://yandex.ru/maps/")
            time.sleep(self.delay * 2)
            
            search_input = self._wait_for_search_input()
            if not search_input:
                self._log("Поисковая строка не найдена")
                return False
            
            if not self._clear_and_type(search_input, city_name):
                return False
            
            time.sleep(0.3)
            search_input.send_keys(Keys.ENTER)
            self._log("Ожидание смены города...")
            time.sleep(2.5)
            
            current_url = self.driver.current_url
            self._log(f"Текущий URL: {current_url}")
            
            search_input = self._wait_for_search_input(timeout=5)
            if search_input:
                self._log(f"Город установлен: {city_name}")
                return True
            else:
                self._log("Поисковая строка не найдена после смены города")
                return False
                
        except Exception as e:
            self._log(f"Ошибка смены города: {str(e)[:80]}")
            return False
    
    def _scroll_results_container(self):
        selectors = [
            "div.scroll__container",
            "div.business-card-list__scroll",
            "div.search-list-view",
            "div.panorama__scroll",
            "div.list-view__scrollable-area",
            "div.search-result-list__scroll",
            "div._scrollable"
        ]
        
        for sel in selectors:
            try:
                container = self.driver.find_element(By.CSS_SELECTOR, sel)
                self.driver.execute_script(
                    "arguments[0].scrollTop = arguments[0].scrollTop + 800", 
                    container
                )
                self._log(f"Скролл через: {sel}")
                return True
            except NoSuchElementException:
                continue
            except Exception as e:
                self._log(f"Ошибка скролла {sel}: {str(e)[:50]}")
                continue
        
        try:
            body = self.driver.find_element(By.TAG_NAME, "body")
            body.send_keys(Keys.PAGE_DOWN)
            self._log("Скролл через PAGE_DOWN")
            return True
        except Exception as e:
            self._log(f"Ошибка скролла PAGE_DOWN: {str(e)[:50]}")
            return False
    
    def search_companies(self, company_type, max_companies=None, city="Ростов-на-Дону"):
        self._log(f"Начинаем поиск: {company_type} в {city}")
        
        if not self.set_city(city):
            self._log("❌ Не удалось установить город, пробуем продолжить...")
        
        search_input = self._wait_for_search_input()
        if not search_input:
            self._log("❌ Поисковая строка не найдена")
            return []
        
        self._log(f"🔍 Вводим запрос: {company_type}")
        if not self._clear_and_type(search_input, company_type):
            self._log("❌ Не удалось ввести запрос")
            return []
        
        time.sleep(0.3)
        search_input.send_keys(Keys.ENTER)
        self._log("⏳ Выполняется поиск...")
        time.sleep(2)
        
        companies_data = []
        seen = set()
        scroll_count, max_scrolls = 0, 50
        last_count, no_change = 0, 0
        
        while scroll_count < max_scrolls:
            try:
                cards = self.driver.find_elements(By.CSS_SELECTOR, "li.search-snippet-view")
                self._log(f"📄 Карточек на странице: {len(cards)}, собрано: {len(companies_data)}")
                
                for card in cards:
                    try:
                        if self._is_ad(card):
                            self._log("⚠️ Пропускаем рекламу")
                            continue
                        
                        # Извлекаем название
                        try:
                            name_el = card.find_element(By.CSS_SELECTOR, "a.link-overlay")
                            name = (name_el.get_attribute('aria-label') or name_el.text).strip()
                        except NoSuchElementException:
                            self._log("⚠️ Не найдено название компании")
                            continue
                        except Exception as e:
                            self._log(f"⚠️ Ошибка названия: {str(e)[:50]}")
                            continue
                        
                        if not name or name in seen:
                            continue
                        seen.add(name)
                        
                        category = "N/A"
                        try:
                            cat = card.find_element(By.CSS_SELECTOR, "div.search-business-snippet-view__categories a")
                            category = cat.text.strip()
                        except:
                            pass
                        
                        address = "N/A"
                        try:
                            addr = card.find_element(By.CSS_SELECTOR, "a.search-business-snippet-view__address")
                            address = addr.text.strip()
                        except:
                            pass
                        
                        yandex_link = self._extract_yandex_link(card, in_card_view=False)
                        
                        data = {
                            'Название': name, 
                            'Категория': category, 
                            'Адрес': address, 
                            'Телефоны': '', 
                            'Сайт': '', 
                            'Рейтинг': '', 
                            'Средняя_оценка': '', 
                            'Отзывов_кол-во': '', 
                            'Отзывы': [], 
                            'Ссылка': yandex_link
                        }
                        
                        try:
                            self.driver.execute_script("arguments[0].click();", name_el)
                            time.sleep(self.delay)
                            
                            if not self._safe_find(By.CSS_SELECTOR, "div.business-card-view", timeout=5):
                                self._log(f"⚠️ Карточка не загрузилась для {name}")
                                body = self.driver.find_element(By.TAG_NAME, "body")
                                body.send_keys(Keys.ESCAPE)
                                time.sleep(0.3)
                                continue
                            
                            time.sleep(0.3)
                            
                            data['Средняя_оценка'] = self._extract_rating()
                            
                            try:
                                rev_count = self.driver.find_element(By.CSS_SELECTOR, "div.business-header-rating-view__text._clickable").text.strip()
                                data['Отзывов_кол-во'] = rev_count
                            except:
                                pass
                            
                            phones = []
                            for pb in self._safe_find_all(By.CSS_SELECTOR, "div.card-phones-view"):
                                try:
                                    exp = pb.find_element(By.CSS_SELECTOR, "div.card-feature-view._interactive")
                                    self.driver.execute_script("arguments[0].click();", exp)
                                    time.sleep(0.15)
                                except:
                                    pass
                                for ph in pb.find_elements(By.CSS_SELECTOR, "div.card-phones-view__phone-number"):
                                    try:
                                        pt = ph.text.strip()
                                        if pt and pt != "Показать телефон":
                                            clean = re.sub(r'[^\d\+\-\(\)\s]', '', pt)
                                            if clean:
                                                phones.append(clean)
                                    except Exception as e:
                                        self._log(f"⚠️ Ошибка телефона: {str(e)[:50]}")
                            data['Телефоны'] = "; ".join(phones) if phones else "Не указан"
                            
                            try:
                                site = self.driver.find_element(By.CSS_SELECTOR, "a.business-urls-view__link")
                                data['Сайт'] = site.get_attribute('href') or site.text.strip()
                            except:
                                data['Сайт'] = "Не указан"
                            
                            if data['Ссылка'] == "N/A":
                                data['Ссылка'] = self._extract_yandex_link(self.driver, in_card_view=True)
                            
                            reviews = self._parse_reviews()
                            data['Отзывы'] = reviews
                            
                            body = self.driver.find_element(By.TAG_NAME, "body")
                            body.send_keys(Keys.ESCAPE)
                            time.sleep(0.3)
                            
                        except Exception as e:
                            self._log(f"Ошибка деталей для {name}: {str(e)[:60]}")
                            try:
                                body = self.driver.find_element(By.TAG_NAME, "body")
                                body.send_keys(Keys.ESCAPE)
                            except:
                                pass
                        
                        companies_data.append(data)
                        self._log(f"{len(companies_data)}. {name} | Оценка: {data['Средняя_оценка']}")
                        
                        if max_companies and len(companies_data) >= max_companies:
                            self._log(f"Лимит {max_companies} достигнут")
                            return companies_data
                        
                        time.sleep(self.delay * 0.5)
                        
                    except StaleElementReferenceException:
                        self._log("StaleElementReferenceException, пропускаем")
                        continue
                    except Exception as e:
                        self._log(f"Непредвиденная ошибка карточки: {str(e)[:60]}")
                        continue
                
                cur = len(companies_data)
                if cur == last_count:
                    no_change += 1
                    self._log(f"ℹНовых нет ({no_change}/3)")
                    if no_change >= 3:
                        self._log("Новые компании не найдены, завершаем")
                        break
                else:
                    no_change, last_count = 0, cur
                
                self._log(f"Прокрутка {scroll_count+1}/{max_scrolls}")
                if not self._scroll_results_container():
                    self._log("Скролл не удался")
                
                try:
                    more = self.driver.find_element(By.CSS_SELECTOR, "button.more-button")
                    if more.is_displayed():
                        self.driver.execute_script("arguments[0].click();", more)
                        self._log("Нажата кнопка 'Показать ещё'")
                        time.sleep(0.5)
                except:
                    pass
                
                scroll_count += 1
                time.sleep(self.delay)
                
            except Exception as e:
                self._log(f"Критическая ошибка в цикле: {str(e)[:80]}")
                break
        
        self._log(f"Всего собрано: {len(companies_data)}")
        return companies_data
    
    def export_to_excel(self, companies, filename):
        if not companies:
            self._log("Нет данных для экспорта")
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Компании"
            
            headers = ['№','Название','Категория','Адрес','Телефоны','Сайт','Рейтинг','Средняя_оценка','Отзывов','Отзывов_текст','Ссылка']
            ws.append(headers)
            
            hf, hf_fill = Font(bold=True,color="FFFFFF"), PatternFill(start_color="366092",end_color="366092",fill_type="solid")
            for c in ws[1]:
                c.font, c.fill, c.alignment = hf, hf_fill, Alignment(horizontal="center")
            
            for i,c in enumerate(companies,1):
                rev_text = "; ".join([f"{r['author']}({r['rating']}):{r['text'][:50]}..." for r in c['Отзывы'][:3]])
                ws.append([i,c['Название'],c['Категория'],c['Адрес'],c['Телефоны'],c['Сайт'],c['Рейтинг'],c['Средняя_оценка'],c['Отзывов_кол-во'],rev_text,c['Ссылка']])
            
            ws_stats = wb.create_sheet("Статистика")
            ws_stats['A1'], ws_stats['A1'].font = "СТАТИСТИКА", Font(bold=True,size=14)
            stats = [["Параметр","Значение"],["Всего",len(companies)],
                    ["С телефонами",sum(1 for x in companies if x['Телефоны']!='Не указан')],
                    ["С сайтом",sum(1 for x in companies if x['Сайт']!='Не указан')],
                    ["С рейтингом",sum(1 for x in companies if x['Рейтинг'])],
                    ["С оценкой",sum(1 for x in companies if x['Средняя_оценка']!='N/A')],
                    ["С отзывами",sum(1 for x in companies if x['Отзывы'])]]
            for row in stats:
                ws_stats.append(row)
            
            for sheet in [ws, ws_stats]:
                for col in sheet.columns:
                    mx = max((len(str(cell.value)) for cell in col if cell.value), default=10)
                    sheet.column_dimensions[col[0].column_letter].width = min(mx+2, 70)
            
            wb.save(filename)
            self._log(f"Excel сохранен: {filename} | Записей: {len(companies)}")
        except Exception as e:
            self._log(f"Ошибка экспорта: {str(e)[:80]}")
    
    def close(self):
        try:
            self.driver.quit()
            self._log("Браузер закрыт")
        except Exception as e:
            self._log(f"Ошибка закрытия: {str(e)[:50]}")


def main():
    parser = YandexMapsParser(headless=False, debug=True)
    try:
        city = input("Город: ").strip()
        ct = input("Тип компании: ").strip()
        lim = input("Лимит (Enter=без): ").strip()
        mx = int(lim) if lim else None
        
        data = parser.search_companies(ct, mx, city)
        if data:
            fn = f"{city.replace(' ','_')}_{ct.replace(' ','_').replace('-','_')}_companies.xlsx"
            parser.export_to_excel(data, fn)
            print(f"\nГотово: {fn}")
        else:
            print("Ничего не найдено")
    except KeyboardInterrupt:
        print("\nПрервано пользователем")
    except Exception as e:
        print(f"Критическая ошибка: {e}")
    finally:
        parser.close()

if __name__ == "__main__":
    main()
